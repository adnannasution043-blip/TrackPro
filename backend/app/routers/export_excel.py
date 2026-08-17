"""
Router Export Excel — generate laporan .xlsx per tahap campaign.

Endpoint:
- GET /laporan-off      : kampanye tahap "off"    → TOTAL OFF + per-tag sheets
- GET /laporan-fix      : kampanye tahap "fix_scale_up" → TOTAL FIX + per-tag sheets
- GET /laporan-harian   : semua tag link → LAP HARIAN (multi-sheet)
"""

import io
from collections import defaultdict
from datetime import date, timedelta
from decimal import Decimal
from uuid import UUID

import sqlalchemy as sa
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.deps import DB, CurrentUser
from app.models.account import MetaAccount, ShopeeAccount
from app.models.campaign import Campaign, CampaignTagMap, TagLink
from app.models.metrics import DailyMetric

router = APIRouter()

_ZERO = Decimal("0")

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

def _hdr_font():   return Font(bold=True, size=10)
def _hdr_fill():   return PatternFill("solid", fgColor="1F4E79")
def _hdr_font_w(): return Font(bold=True, color="FFFFFF", size=10)
def _total_fill(): return PatternFill("solid", fgColor="D6E4F0")
def _total_font(): return Font(bold=True, size=10)


def _thin_border():
    s = Side(style="thin", color="C0C0C0")
    return Border(left=s, right=s, top=s, bottom=s)


def _apply_header(ws, cols: list[str], row: int = 1):
    for c, label in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = _hdr_font_w()
        cell.fill = _hdr_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()


def _apply_total_row(ws, row: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _total_font()
        cell.fill = _total_fill()
        cell.border = _thin_border()


# ---------------------------------------------------------------------------
# Shared helper: build per-tahap workbook (OFF / FIX / dll)
# ---------------------------------------------------------------------------

async def _build_per_tahap_wb(
    tahap: str,
    total_sheet_name: str,
    user_id,
    db,
    dates: list[date],
    meta_account_id=None,
) -> Workbook:
    tahap_label = {
        "off": "OFF", "fix_scale_up": "FIX",
        "filter": "FILTER", "pra_filter": "PRA FILTER",
    }.get(tahap, tahap.upper())

    tanggal_dari, tanggal_sampai = dates[0], dates[-1]

    q_camp = (
        sa.select(Campaign)
        .join(MetaAccount, Campaign.meta_account_id == MetaAccount.id)
        .where(MetaAccount.user_id == user_id, Campaign.tahap == tahap)
    )
    if meta_account_id:
        q_camp = q_camp.where(Campaign.meta_account_id == meta_account_id)
    campaigns = (await db.execute(q_camp)).scalars().all()

    if not campaigns:
        raise HTTPException(404, f"Tidak ada campaign dengan tahap {tahap_label} untuk periode ini.")

    camp_ids = [c.id for c in campaigns]

    maps = (await db.execute(
        sa.select(CampaignTagMap).where(CampaignTagMap.campaign_id.in_(camp_ids))
    )).scalars().all()

    tag_to_camp: dict[UUID, UUID] = {}
    for m in maps:
        tag_to_camp[m.tag_link_id] = m.campaign_id
    tag_ids = list(tag_to_camp.keys())

    tag_objs = (await db.execute(
        sa.select(TagLink).where(TagLink.id.in_(tag_ids))
    )).scalars().all() if tag_ids else []
    tag_name: dict[UUID, str] = {t.id: t.tag for t in tag_objs}

    meta_rows = (await db.execute(
        sa.select(DailyMetric)
        .where(
            DailyMetric.campaign_id.in_(camp_ids),
            DailyMetric.tanggal >= tanggal_dari,
            DailyMetric.tanggal <= tanggal_sampai,
        )
    )).scalars().all()

    meta_by: dict[UUID, dict[date, dict]] = defaultdict(dict)
    for r in meta_rows:
        meta_by[r.campaign_id][r.tanggal] = {
            "spend": r.spend_idr or _ZERO,
            "clicks_meta": r.clicks_meta or 0,
        }

    shopee_rows = (await db.execute(
        sa.select(DailyMetric)
        .where(
            DailyMetric.tag_link_id.in_(tag_ids),
            DailyMetric.tanggal >= tanggal_dari,
            DailyMetric.tanggal <= tanggal_sampai,
        )
    )).scalars().all() if tag_ids else []

    shopee_by: dict[UUID, dict[date, dict]] = defaultdict(dict)
    for r in shopee_rows:
        shopee_by[r.tag_link_id][r.tanggal] = {
            "komisi": r.commission_idr or _ZERO,
            "clicks_shopee": r.clicks_shopee or 0,
        }

    wb = Workbook()
    wb.remove(wb.active)

    COLS = ["TGL", "Spend", "(+) 5%", "KOMISI", "PROFIT", "% PROFIT",
            "Klik FP", "Klik Shp", "(%) Klik", "CPC FP", "CPC Shp", "Tag Link", "Status", "Note"]
    N = len(COLS)

    total_by_date: dict[date, dict] = defaultdict(lambda: {
        "spend": _ZERO, "komisi": _ZERO, "clicks_meta": 0, "clicks_shopee": 0
    })

    sheets_order = [("tag", tid) for tid in tag_ids]
    unmapped = [c for c in campaigns if c.id not in {tag_to_camp[t] for t in tag_ids}]
    sheets_order += [("camp", c.id) for c in unmapped]

    for kind, eid in sheets_order:
        if kind == "tag":
            tag_id = eid
            camp_id = tag_to_camp[tag_id]
            sheet_name = tag_name.get(tag_id, str(tag_id)[:30])[:31]
        else:
            camp_id = eid
            tag_id = None
            camp = next((c for c in campaigns if c.id == camp_id), None)
            sheet_name = (camp.nama_campaign[:31] if camp else str(camp_id)[:31])

        ws = wb.create_sheet(title=sheet_name)
        _apply_header(ws, COLS)
        data_rows = len(dates)

        for i, d in enumerate(dates, start=2):
            m = meta_by[camp_id].get(d, {}) if camp_id else {}
            s = shopee_by[tag_id].get(d, {}) if tag_id else {}
            spend = float(m.get("spend", 0))
            komisi = float(s.get("komisi", 0))
            clicks_meta = m.get("clicks_meta", 0)
            clicks_shopee = s.get("clicks_shopee", 0)

            ws.cell(row=i, column=1, value=d.strftime("%Y-%m-%d"))
            ws.cell(row=i, column=2, value=spend or None)
            ws.cell(row=i, column=3, value=f"=B{i}+(B{i}*5%)" if spend else None)
            ws.cell(row=i, column=4, value=komisi or None)
            ws.cell(row=i, column=5, value=f"=D{i}-C{i}")
            ws.cell(row=i, column=6, value=f"=IF(C{i}=0,\"\",E{i}/C{i})")
            ws.cell(row=i, column=7, value=clicks_meta or None)
            ws.cell(row=i, column=8, value=clicks_shopee or None)
            ws.cell(row=i, column=9, value=f"=IF(G{i}=0,\"\",H{i}/G{i})")
            ws.cell(row=i, column=10, value=f"=IF(G{i}=0,\"\",C{i}/G{i})")
            ws.cell(row=i, column=11, value=f"=IF(H{i}=0,\"\",D{i}/H{i})")
            ws.cell(row=i, column=12, value=tag_name.get(tag_id, "") if tag_id else "")
            ws.cell(row=i, column=6).number_format = "0.00%"
            ws.cell(row=i, column=9).number_format = "0.00%"

            if tag_id:
                td = total_by_date[d]
                td["spend"] += Decimal(str(spend))
                td["komisi"] += Decimal(str(komisi))
                td["clicks_meta"] += clicks_meta
                td["clicks_shopee"] += clicks_shopee

        tr = data_rows + 2
        ws.cell(row=tr, column=1, value="TOTAL")
        ws.cell(row=tr, column=2, value=f"=SUM(B2:B{data_rows + 1})")
        ws.cell(row=tr, column=3, value=f"=B{tr}+(B{tr}*5%)")
        ws.cell(row=tr, column=4, value=f"=SUM(D2:D{data_rows + 1})")
        ws.cell(row=tr, column=5, value=f"=D{tr}-C{tr}")
        ws.cell(row=tr, column=6, value=f"=IF(C{tr}=0,\"\",E{tr}/C{tr})")
        ws.cell(row=tr, column=7, value=f"=SUM(G2:G{data_rows + 1})")
        ws.cell(row=tr, column=8, value=f"=SUM(H2:H{data_rows + 1})")
        ws.cell(row=tr, column=9, value=f"=IF(G{tr}=0,\"\",H{tr}/G{tr})")
        ws.cell(row=tr, column=10, value=f"=IF(G{tr}=0,\"\",C{tr}/G{tr})")
        ws.cell(row=tr, column=11, value=f"=IF(H{tr}=0,\"\",D{tr}/H{tr})")
        _apply_total_row(ws, tr, N)
        ws.cell(row=tr, column=6).number_format = "0.00%"
        ws.cell(row=tr, column=9).number_format = "0.00%"

        for col, w in zip("ABCDEFGHIJKLMN", [12,14,14,14,14,10,10,10,10,12,12,22,10,20]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "B2"

    # Summary sheet di index 0
    TCOLS = ["TGL", "Spend", "(+) 5%", "KOMISI", "PROFIT", "% PROFIT",
             "Klik FP", "Klik Shp", "(%) Klik", "CPC FP", "CPC Shp"]
    ws_t = wb.create_sheet(title=total_sheet_name, index=0)
    _apply_header(ws_t, TCOLS)

    for i, d in enumerate(dates, start=2):
        td = total_by_date[d]
        spend = float(td["spend"])
        komisi = float(td["komisi"])
        cm = td["clicks_meta"]
        cs = td["clicks_shopee"]
        ws_t.cell(row=i, column=1, value=d.strftime("%Y-%m-%d"))
        ws_t.cell(row=i, column=2, value=spend or None)
        ws_t.cell(row=i, column=3, value=f"=B{i}+(B{i}*5%)" if spend else None)
        ws_t.cell(row=i, column=4, value=komisi or None)
        ws_t.cell(row=i, column=5, value=f"=D{i}-C{i}")
        ws_t.cell(row=i, column=6, value=f"=IF(C{i}=0,\"\",E{i}/C{i})")
        ws_t.cell(row=i, column=7, value=cm or None)
        ws_t.cell(row=i, column=8, value=cs or None)
        ws_t.cell(row=i, column=9, value=f"=IF(G{i}=0,\"\",H{i}/G{i})")
        ws_t.cell(row=i, column=10, value=f"=IF(G{i}=0,\"\",C{i}/G{i})")
        ws_t.cell(row=i, column=11, value=f"=IF(H{i}=0,\"\",D{i}/H{i})")
        ws_t.cell(row=i, column=6).number_format = "0.00%"
        ws_t.cell(row=i, column=9).number_format = "0.00%"

    nd = len(dates)
    tr = nd + 2
    ws_t.cell(row=tr, column=1, value="TOTAL")
    ws_t.cell(row=tr, column=2, value=f"=SUM(B2:B{nd + 1})")
    ws_t.cell(row=tr, column=3, value=f"=B{tr}+(B{tr}*5%)")
    ws_t.cell(row=tr, column=4, value=f"=SUM(D2:D{nd + 1})")
    ws_t.cell(row=tr, column=5, value=f"=D{tr}-C{tr}")
    ws_t.cell(row=tr, column=6, value=f"=IF(C{tr}=0,\"\",E{tr}/C{tr})")
    ws_t.cell(row=tr, column=7, value=f"=SUM(G2:G{nd + 1})")
    ws_t.cell(row=tr, column=8, value=f"=SUM(H2:H{nd + 1})")
    ws_t.cell(row=tr, column=9, value=f"=IF(G{tr}=0,\"\",H{tr}/G{tr})")
    ws_t.cell(row=tr, column=10, value=f"=IF(G{tr}=0,\"\",C{tr}/G{tr})")
    ws_t.cell(row=tr, column=11, value=f"=IF(H{tr}=0,\"\",D{tr}/H{tr})")
    _apply_total_row(ws_t, tr, len(TCOLS))
    ws_t.cell(row=tr, column=6).number_format = "0.00%"
    ws_t.cell(row=tr, column=9).number_format = "0.00%"
    for col, w in zip("ABCDEFGHIJK", [12,14,14,14,14,10,10,10,10,12,12]):
        ws_t.column_dimensions[col].width = w
    ws_t.freeze_panes = "B2"

    return wb


def _stream_wb(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Endpoints: OFF & FIX
# ---------------------------------------------------------------------------

@router.get("/laporan-off")
async def export_laporan_off(
    current_user: CurrentUser,
    db: DB,
    tanggal_dari: date = Query(...),
    tanggal_sampai: date = Query(...),
    meta_account_id: UUID | None = Query(None),
    shopee_account_id: UUID | None = Query(None),
):
    dates = [tanggal_dari + timedelta(days=i)
             for i in range((tanggal_sampai - tanggal_dari).days + 1)]
    wb = await _build_per_tahap_wb("off", "TOTAL OFF", current_user.id, db, dates, meta_account_id)
    bulan = tanggal_dari.strftime("%B %Y").upper()
    return _stream_wb(wb, f"OFF FILTER META {bulan}.xlsx")


@router.get("/laporan-fix")
async def export_laporan_fix(
    current_user: CurrentUser,
    db: DB,
    tanggal_dari: date = Query(...),
    tanggal_sampai: date = Query(...),
    meta_account_id: UUID | None = Query(None),
    shopee_account_id: UUID | None = Query(None),
):
    dates = [tanggal_dari + timedelta(days=i)
             for i in range((tanggal_sampai - tanggal_dari).days + 1)]
    wb = await _build_per_tahap_wb("fix_scale_up", "TOTAL FIX", current_user.id, db, dates, meta_account_id)
    bulan = tanggal_dari.strftime("%B %Y").upper()
    return _stream_wb(wb, f"FIX META {bulan}.xlsx")


# ---------------------------------------------------------------------------
# Endpoint: PRA FILTER ADV — per campaign pra_filter, dua snapshot tanggal
# ---------------------------------------------------------------------------

@router.get("/laporan-pra-filter")
async def export_laporan_pra_filter(
    current_user: CurrentUser,
    db: DB,
    tanggal_dari: date = Query(...),
    tanggal_sampai: date = Query(...),
    meta_account_id: UUID | None = Query(None),
    shopee_account_id: UUID | None = Query(None),
):
    # Ambil campaign pra_filter milik user
    q_camp = (
        sa.select(Campaign)
        .join(MetaAccount, Campaign.meta_account_id == MetaAccount.id)
        .where(MetaAccount.user_id == current_user.id, Campaign.tahap == "pra_filter")
    )
    if meta_account_id:
        q_camp = q_camp.where(Campaign.meta_account_id == meta_account_id)
    campaigns = (await db.execute(q_camp)).scalars().all()
    if not campaigns:
        raise HTTPException(404, "Tidak ada campaign dengan tahap Pra Filter.")

    camp_ids = [c.id for c in campaigns]

    # Tag link mapping
    maps = (await db.execute(
        sa.select(CampaignTagMap).where(CampaignTagMap.campaign_id.in_(camp_ids))
    )).scalars().all()
    camp_to_tags: dict[UUID, list[UUID]] = defaultdict(list)
    for m in maps:
        camp_to_tags[m.campaign_id].append(m.tag_link_id)
    all_tag_ids = list({t for tags in camp_to_tags.values() for t in tags})

    # Spend + klik per (campaign, tanggal) — snapshot hari pertama dan terakhir
    meta_rows = (await db.execute(
        sa.select(DailyMetric.campaign_id, DailyMetric.tanggal,
                  DailyMetric.spend_idr, DailyMetric.clicks_meta)
        .where(
            DailyMetric.campaign_id.in_(camp_ids),
            DailyMetric.tanggal.in_([tanggal_dari, tanggal_sampai]),
        )
    )).all()
    meta_by: dict[UUID, dict[date, dict]] = defaultdict(dict)
    for r in meta_rows:
        meta_by[r.campaign_id][r.tanggal] = {
            "spend": float(r.spend_idr or 0),
            "clicks": r.clicks_meta or 0,
        }

    # Komisi + klik shopee per (tag_link) — total seluruh periode
    shopee_rows = (await db.execute(
        sa.select(DailyMetric.tag_link_id,
                  sa.func.sum(DailyMetric.commission_idr).label("total_komisi"),
                  sa.func.sum(DailyMetric.clicks_shopee).label("total_klik"))
        .where(
            DailyMetric.tag_link_id.in_(all_tag_ids) if all_tag_ids else sa.false(),
            DailyMetric.tanggal >= tanggal_dari,
            DailyMetric.tanggal <= tanggal_sampai,
        )
        .group_by(DailyMetric.tag_link_id)
    )).all() if all_tag_ids else []

    komisi_by_tag: dict[UUID, float] = {}
    klik_by_tag: dict[UUID, int] = {}
    for r in shopee_rows:
        komisi_by_tag[r.tag_link_id] = float(r.total_komisi or 0)
        klik_by_tag[r.tag_link_id] = int(r.total_klik or 0)

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)

    d1_label = tanggal_dari.strftime("%-d %B - %H:%M").upper() if hasattr(tanggal_dari, 'strftime') else tanggal_dari.strftime("%d %B").upper()
    d2_label = tanggal_sampai.strftime("%-d %B - %H:%M").upper() if hasattr(tanggal_sampai, 'strftime') else tanggal_sampai.strftime("%d %B").upper()
    # Windows-safe date label
    d1_label = tanggal_dari.strftime("%d %B").lstrip("0").upper()
    d2_label = tanggal_sampai.strftime("%d %B").lstrip("0").upper()

    sheet_name = tanggal_dari.strftime("%d %B ADV").lstrip("0").upper()[:31]
    ws = wb.create_sheet(sheet_name)

    # ── Row 1: headers utama ──────────────────────────────────────────────
    def _h(cell, value):
        cell.value = value
        cell.font = _hdr_font_w()
        cell.fill = _hdr_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()

    _h(ws["A1"], "NAMA CAMPAIGN")
    _h(ws["B1"], "JENIS")
    _h(ws["C1"], d1_label)    # span C-E
    _h(ws["G1"], d2_label)    # span G-J
    _h(ws["K1"], "KOMISI")
    _h(ws["L1"], "(%) PROFIT")
    _h(ws["M1"], "KLIK SHP")
    _h(ws["N1"], "(%) KLIK")
    _h(ws["O1"], "NOTE")
    ws.merge_cells("C1:E1")
    ws.merge_cells("G1:J1")

    # ── Row 2: sub-headers ────────────────────────────────────────────────
    tayang_label = f"TAYANG {d1_label}"
    _h(ws["A2"], tayang_label)
    _h(ws["C2"], "CPC")
    _h(ws["D2"], "SPENT")
    _h(ws["E2"], "KLIK META")
    _h(ws["G2"], "CPC")
    _h(ws["H2"], "SPENT")
    _h(ws["I2"], "(+) 5%")
    _h(ws["J2"], "KLIK META")

    # ── Data rows ─────────────────────────────────────────────────────────
    FMT_IDR = "#,##0"
    FMT_PCT = "0.00%"

    for ridx, camp in enumerate(sorted(campaigns, key=lambda c: c.nama_campaign), start=3):
        snap1 = meta_by[camp.id].get(tanggal_dari, {})
        snap2 = meta_by[camp.id].get(tanggal_sampai, {})
        tag_ids_for_camp = camp_to_tags.get(camp.id, [])
        total_komisi = sum(komisi_by_tag.get(t, 0) for t in tag_ids_for_camp)
        total_klik_shp = sum(klik_by_tag.get(t, 0) for t in tag_ids_for_camp)

        ws.cell(ridx, 1, camp.nama_campaign)
        ws.cell(ridx, 2, "")  # JENIS — isi manual

        # Snapshot 1
        s1_spend = snap1.get("spend", 0)
        s1_klik = snap1.get("clicks", 0)
        if s1_spend:
            ws.cell(ridx, 4, s1_spend).number_format = FMT_IDR
        if s1_klik:
            ws.cell(ridx, 5, s1_klik)
        if s1_spend and s1_klik:
            ws.cell(ridx, 3, f"=D{ridx}/E{ridx}").number_format = FMT_IDR

        # Snapshot 2
        s2_spend = snap2.get("spend", 0)
        s2_klik = snap2.get("clicks", 0)
        if s2_spend:
            ws.cell(ridx, 8, s2_spend).number_format = FMT_IDR
            ws.cell(ridx, 9, f"=H{ridx}+(H{ridx}*5%)").number_format = FMT_IDR
        if s2_klik:
            ws.cell(ridx, 10, s2_klik)
        if s2_spend and s2_klik:
            ws.cell(ridx, 7, f"=H{ridx}/J{ridx}").number_format = FMT_IDR

        # Komisi & derived
        if total_komisi:
            ws.cell(ridx, 11, total_komisi).number_format = FMT_IDR
        ws.cell(ridx, 12, f"=IF(I{ridx}=0,\"\",(K{ridx}-I{ridx})/I{ridx})").number_format = FMT_PCT
        if total_klik_shp:
            ws.cell(ridx, 13, total_klik_shp)
        ws.cell(ridx, 14, f"=IF(J{ridx}=0,\"\",M{ridx}/J{ridx})").number_format = FMT_PCT

    # Lebar kolom
    for col, w in zip("ABCDEFGHIJKLMNO",
                      [32, 10, 12, 14, 12, 4, 12, 14, 14, 12, 14, 12, 12, 10, 20]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "C3"

    bulan = tanggal_dari.strftime("%B %Y").upper()
    return _stream_wb(wb, f"PRA FILTER ADV {bulan}.xlsx")


# ---------------------------------------------------------------------------
# Endpoint: LAP HARIAN — semua tag link, tanpa filter tahap
# ---------------------------------------------------------------------------

@router.get("/laporan-harian")
async def export_laporan_harian(
    current_user: CurrentUser,
    db: DB,
    tanggal_dari: date = Query(...),
    tanggal_sampai: date = Query(...),
    meta_account_id: UUID | None = Query(None),
):
    dates = [tanggal_dari + timedelta(days=i)
             for i in range((tanggal_sampai - tanggal_dari).days + 1)]
    DATA_START = 8
    last_data_row = DATA_START + len(dates) - 1

    # 1. Tag link milik user ini
    q_tags = (
        sa.select(TagLink)
        .join(ShopeeAccount, TagLink.shopee_account_id == ShopeeAccount.id)
        .where(ShopeeAccount.user_id == current_user.id)
    )
    tag_objs = (await db.execute(q_tags)).scalars().all()
    if not tag_objs:
        raise HTTPException(404, "Tidak ada tag link yang ditemukan.")

    tag_ids = [t.id for t in tag_objs]
    tag_name_map: dict[UUID, str] = {t.id: t.tag for t in tag_objs}

    # 2. Semua campaign milik user
    q_camp = (
        sa.select(Campaign)
        .join(MetaAccount, Campaign.meta_account_id == MetaAccount.id)
        .where(MetaAccount.user_id == current_user.id)
    )
    if meta_account_id:
        q_camp = q_camp.where(Campaign.meta_account_id == meta_account_id)
    campaigns = (await db.execute(q_camp)).scalars().all()
    camp_ids = [c.id for c in campaigns]

    # 3. Campaign → tag_link mapping
    maps = (await db.execute(
        sa.select(CampaignTagMap).where(CampaignTagMap.campaign_id.in_(camp_ids))
    )).scalars().all() if camp_ids else []

    tag_to_camps: dict[UUID, list[UUID]] = defaultdict(list)
    for m in maps:
        if m.tag_link_id in tag_name_map:
            tag_to_camps[m.tag_link_id].append(m.campaign_id)

    # 4. Meta spend per (campaign, tanggal)
    meta_rows = (await db.execute(
        sa.select(DailyMetric.campaign_id, DailyMetric.tanggal, DailyMetric.spend_idr)
        .where(
            DailyMetric.campaign_id.in_(camp_ids),
            DailyMetric.tanggal >= tanggal_dari,
            DailyMetric.tanggal <= tanggal_sampai,
            DailyMetric.spend_idr.isnot(None),
        )
    )).all() if camp_ids else []

    spend_by: dict[UUID, dict[date, Decimal]] = defaultdict(lambda: defaultdict(lambda: _ZERO))
    for r in meta_rows:
        spend_by[r.campaign_id][r.tanggal] += r.spend_idr or _ZERO

    # 5. Shopee commission per (tag_link, tanggal)
    shopee_rows = (await db.execute(
        sa.select(DailyMetric.tag_link_id, DailyMetric.tanggal, DailyMetric.commission_idr)
        .where(
            DailyMetric.tag_link_id.in_(tag_ids),
            DailyMetric.tanggal >= tanggal_dari,
            DailyMetric.tanggal <= tanggal_sampai,
            DailyMetric.commission_idr.isnot(None),
        )
    )).all()

    komisi_by: dict[UUID, dict[date, Decimal]] = defaultdict(lambda: defaultdict(lambda: _ZERO))
    for r in shopee_rows:
        komisi_by[r.tag_link_id][r.tanggal] += r.commission_idr or _ZERO

    # 6. Build workbook
    wb = Workbook()
    wb.remove(wb.active)

    def _fmt_dates(ws, col="B"):
        for idx, d in enumerate(dates):
            r = DATA_START + idx
            ws[f"{col}{r}"] = d
            ws[f"{col}{r}"].number_format = "DD/MM/YYYY"

    # ── Sheet: % PAJAK ────────────────────────────────────────────────────────
    ws_pajak = wb.create_sheet("% PAJAK")
    for label, col in [("TGL", "B"), ("% PAJAK", "C"), ("Biaya Layanan", "D")]:
        cell = ws_pajak[f"{col}4"]
        cell.value = label
        cell.font = _hdr_font_w()
        cell.fill = _hdr_fill()
        cell.alignment = Alignment(horizontal="center")
    _fmt_dates(ws_pajak, "B")
    ws_pajak.column_dimensions["B"].width = 14
    ws_pajak.column_dimensions["C"].width = 12
    ws_pajak.column_dimensions["D"].width = 16

    # ── Sheet: LIVE GILANG (blank template) ───────────────────────────────────
    ws_live = wb.create_sheet("LIVE GILANG")
    for label, col in [("TGL", "B"), ("Komisi Bersih before Tax", "E"), ("Komisi Bersih after Tax", "G")]:
        c = ws_live[f"{col}4"]
        c.value = label
        c.font = _hdr_font_w()
        c.fill = _hdr_fill()
    _fmt_dates(ws_live, "B")
    ws_live.column_dimensions["B"].width = 14

    # ── Sheet: Organik FP (blank template) ────────────────────────────────────
    ws_org = wb.create_sheet("Organik FP")
    for label, col in [("TGL", "B"), ("Komisi Kotor Harian Fanpage", "C"),
                        ("Komisi Bersih before Tax", "L"), ("Komisi Bersih after Tax", "P")]:
        c = ws_org[f"{col}4"]
        c.value = label
        c.font = _hdr_font_w()
        c.fill = _hdr_fill()
    _fmt_dates(ws_org, "B")
    ws_org.column_dimensions["B"].width = 14

    # ── Sheet: TOTAL ORGANIK ─────────────────────────────────────────────────
    ws_torg = wb.create_sheet("TOTAL ORGANIK")
    for label, col in [("TGL", "B"), ("Komisi Bersih after Tax", "C")]:
        c = ws_torg[f"{col}4"]
        c.value = label
        c.font = _hdr_font_w()
        c.fill = _hdr_fill()
    ws_torg["C6"] = "Komisi FP"
    ws_torg["D6"] = "TOTAL KOMISI"
    _fmt_dates(ws_torg, "B")
    for idx in range(len(dates)):
        r = DATA_START + idx
        ws_torg[f"C{r}"] = f"='Organik FP'!R{r}"
        ws_torg[f"D{r}"] = f"=SUM(C{r})"
    ws_torg.column_dimensions["B"].width = 14
    ws_torg.column_dimensions["C"].width = 18
    ws_torg.column_dimensions["D"].width = 16

    # ── Sheet per tag link ────────────────────────────────────────────────────
    tag_sheet_names: list[str] = []
    for tag_id in tag_ids:
        tag_n = tag_name_map[tag_id]
        sheet_name = tag_n[:31]
        tag_sheet_names.append(sheet_name)
        ws = wb.create_sheet(sheet_name)

        headers = {
            "C": "TGL", "D": "Budget Iklan", "E": "(+) 5%",
            "G": f"Komisi harian {tag_n[:15]}",
            "H": "Profit kotor Iklan",
            "J": "Komisi Bersih before Tax",
            "K": "Komisi Bersih Iklan after Tax",
            "L": "Profit bersih Iklan",
        }
        for col, label in headers.items():
            cell = ws[f"{col}5"]
            cell.value = label
            cell.font = _hdr_font_w()
            cell.fill = _hdr_fill()
            cell.alignment = Alignment(horizontal="center")

        linked_camps = tag_to_camps.get(tag_id, [])

        for idx, d in enumerate(dates):
            r = DATA_START + idx
            spend = sum(float(spend_by[cid].get(d, _ZERO)) for cid in linked_camps)
            komisi = float(komisi_by[tag_id].get(d, _ZERO))

            ws[f"C{r}"] = d
            ws[f"C{r}"].number_format = "DD/MM/YYYY"
            if spend:
                ws[f"D{r}"] = spend
                ws[f"D{r}"].number_format = "#,##0"
                ws[f"E{r}"] = f"=D{r}+(D{r}*5%)"
                ws[f"E{r}"].number_format = "#,##0"
            if komisi:
                ws[f"G{r}"] = komisi
                ws[f"G{r}"].number_format = "#,##0"
            ws[f"H{r}"] = f"=IF(OR(G{r}=\"\",E{r}=\"\"),\"\",G{r}-E{r})"
            ws[f"H{r}"].number_format = "#,##0"
            ws[f"J{r}"] = f"=IF(G{r}=\"\",\"\",G{r}-'% PAJAK'!D{r})"
            ws[f"J{r}"].number_format = "#,##0"
            ws[f"K{r}"] = f"=IF(J{r}=\"\",\"\",J{r}-(J{r}*'% PAJAK'!C{r}))"
            ws[f"K{r}"].number_format = "#,##0"
            ws[f"L{r}"] = f"=IF(OR(K{r}=\"\",E{r}=\"\"),\"\",K{r}-E{r})"
            ws[f"L{r}"].number_format = "#,##0"

        tr = last_data_row + 1
        ws[f"C{tr}"] = "TOTAL"
        ws[f"C{tr}"].font = _total_font()
        for col, formula in [
            ("D", f"=SUM(D{DATA_START}:D{last_data_row})"),
            ("E", f"=D{tr}+(D{tr}*5%)"),
            ("G", f"=SUM(G{DATA_START}:G{last_data_row})"),
            ("H", f"=G{tr}-E{tr}"),
            ("J", f"=SUM(J{DATA_START}:J{last_data_row})"),
            ("K", f"=SUM(K{DATA_START}:K{last_data_row})"),
            ("L", f"=K{tr}-E{tr}"),
        ]:
            ws[f"{col}{tr}"] = formula
            ws[f"{col}{tr}"].number_format = "#,##0"
        _apply_total_row(ws, tr, 12)

        for col, w in zip("CDEFGHJKL", [12, 14, 14, 4, 16, 14, 4, 16, 16, 14]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "D6"

    # ── Sheet: TOTAL IKLAN ────────────────────────────────────────────────────
    ws_ti = wb.create_sheet("TOTAL IKLAN")
    ti_headers = {
        "C": "TGL", "D": "Total Budget Iklan", "F": "Total Komisi harian Iklan",
        "G": "Profit kotor Iklan", "I": "Komisi Bersih before Tax",
        "J": "Komisi Bersih Iklan after Tax", "K": "Profit bersih Iklan",
    }
    for col, label in ti_headers.items():
        cell = ws_ti[f"{col}5"]
        cell.value = label
        cell.font = _hdr_font_w()
        cell.fill = _hdr_fill()
        cell.alignment = Alignment(horizontal="center")

    def _tag_sum(col, row):
        if not tag_sheet_names:
            return 0
        return "=" + "+".join(f"'{s}'!{col}{row}" for s in tag_sheet_names)

    for idx, d in enumerate(dates):
        r = DATA_START + idx
        ws_ti[f"C{r}"] = d
        ws_ti[f"C{r}"].number_format = "DD/MM/YYYY"
        for col, src_col in [("D", "D"), ("F", "G"), ("I", "J")]:
            ws_ti[f"{col}{r}"] = _tag_sum(src_col, r)
            ws_ti[f"{col}{r}"].number_format = "#,##0"
        ws_ti[f"G{r}"] = f"=F{r}-D{r}"
        ws_ti[f"G{r}"].number_format = "#,##0"
        ws_ti[f"J{r}"] = f"=I{r}-(I{r}*'% PAJAK'!C{r})"
        ws_ti[f"J{r}"].number_format = "#,##0"
        ws_ti[f"K{r}"] = f"=J{r}-D{r}"
        ws_ti[f"K{r}"].number_format = "#,##0"

    tr = last_data_row + 1
    ws_ti[f"C{tr}"] = "TOTAL"
    for col in ["D", "F", "G", "I", "J", "K"]:
        ws_ti[f"{col}{tr}"] = f"=SUM({col}{DATA_START}:{col}{last_data_row})"
        ws_ti[f"{col}{tr}"].number_format = "#,##0"
    _apply_total_row(ws_ti, tr, 11)
    for col, w in zip("CDEFGIJK", [12, 14, 4, 16, 14, 4, 16, 16, 14]):
        ws_ti.column_dimensions[col].width = w
    ws_ti.freeze_panes = "D6"

    # ── Sheet: TOTAL KOMISI (partial) ─────────────────────────────────────────
    ws_tk = wb.create_sheet("TOTAL KOMISI")
    for label, col in [("TGL", "B"), ("Komisi Bersih after Tax", "C"),
                        ("Komisi Iklan", "E"), ("Total Komisi Masuk", "F"),
                        ("Budget Iklan", "I"), ("Profit Iklan", "J")]:
        c = ws_tk[f"{col}4"]
        c.value = label
        c.font = _hdr_font_w()
        c.fill = _hdr_fill()
    ws_tk["C6"] = "Komisi LIVE"
    ws_tk["D6"] = "Komisi Organik"
    ws_tk["E6"] = "Komisi Iklan"
    ws_tk["F6"] = "Total Komisi Masuk"
    ws_tk["I6"] = "Budget Iklan"
    ws_tk["J6"] = "Profit Iklan"
    for idx, d in enumerate(dates):
        r = DATA_START + idx
        ws_tk[f"B{r}"] = d
        ws_tk[f"B{r}"].number_format = "DD/MM/YYYY"
        ws_tk[f"E{r}"] = f"='TOTAL IKLAN'!J{r}"
        ws_tk[f"E{r}"].number_format = "#,##0"
        ws_tk[f"F{r}"] = f"=SUM(C{r}:E{r})"
        ws_tk[f"F{r}"].number_format = "#,##0"
        ws_tk[f"I{r}"] = f"='TOTAL IKLAN'!D{r}"
        ws_tk[f"I{r}"].number_format = "#,##0"
        ws_tk[f"J{r}"] = f"=E{r}-I{r}"
        ws_tk[f"J{r}"].number_format = "#,##0"
    ws_tk.column_dimensions["B"].width = 14
    ws_tk.freeze_panes = "C7"

    # ── Sheet: komisi kotor (partial) ─────────────────────────────────────────
    ws_kk = wb.create_sheet("komisi kotor")
    for label, col in [("TGL", "B"), ("Komisi Live", "C"), ("Komisi Story", "D"),
                        ("Komisi Few Feed", "E"), ("Total Komisi FP", "F")]:
        c = ws_kk[f"{col}4"]
        c.value = label
        c.font = _hdr_font_w()
        c.fill = _hdr_fill()

    tag_kk_cols = []
    start_col_idx = 7
    for i, (tag_id, sname) in enumerate(zip(tag_ids, tag_sheet_names)):
        col_letter = get_column_letter(start_col_idx + i)
        tag_kk_cols.append((col_letter, sname))
        ws_kk[f"{col_letter}6"] = f"Komisi {sname[:12]}"
        ws_kk.column_dimensions[col_letter].width = 16

    total_iklan_col = get_column_letter(start_col_idx + len(tag_ids))
    total_kotor_col = get_column_letter(start_col_idx + len(tag_ids) + 1)
    for col, label in [(total_iklan_col, "Total Komisi Iklan"), (total_kotor_col, "Total Komisi Kotor")]:
        c = ws_kk[f"{col}4"]
        c.value = label
        c.font = _hdr_font_w()
        c.fill = _hdr_fill()
        ws_kk.column_dimensions[col].width = 18

    for idx, d in enumerate(dates):
        r = DATA_START + idx
        ws_kk[f"B{r}"] = d
        ws_kk[f"B{r}"].number_format = "DD/MM/YYYY"
        for col_letter, sname in tag_kk_cols:
            ws_kk[f"{col_letter}{r}"] = f"='{sname}'!G{r}"
            ws_kk[f"{col_letter}{r}"].number_format = "#,##0"
        if tag_kk_cols:
            ws_kk[f"{total_iklan_col}{r}"] = "=" + "+".join(f"{cl}{r}" for cl, _ in tag_kk_cols)
        else:
            ws_kk[f"{total_iklan_col}{r}"] = 0
        ws_kk[f"{total_iklan_col}{r}"].number_format = "#,##0"
        ws_kk[f"{total_kotor_col}{r}"] = f"=F{r}+{total_iklan_col}{r}"
        ws_kk[f"{total_kotor_col}{r}"].number_format = "#,##0"
    ws_kk.column_dimensions["B"].width = 14
    ws_kk.freeze_panes = "C7"

    bulan = tanggal_dari.strftime("%B %Y").upper()
    return _stream_wb(wb, f"LAP HARIAN {bulan}.xlsx")
