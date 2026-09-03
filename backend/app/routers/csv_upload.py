"""
Router upload CSV — 3 jenis file:
1. Meta Ads        → daily_metrics (campaign_id, spend_idr, clicks_meta)
2. Shopee Commission → daily_metrics (tag_link_id, orders_*, commission_idr, sales_idr)
                       + order_snapshots (per order_id, append-only)
3. Shopee Click    → daily_metrics (tag_link_id, clicks_shopee)

Upsert strategy:
- Tiap jenis CSV hanya update kolom miliknya sendiri, kolom lain tidak disentuh.
  Ini agar upload Meta dan upload Shopee untuk tanggal yang sama tidak saling
  menimpa (schema mengizinkan satu baris DailyMetric per (campaign_id, date)
  dan satu per (tag_link_id, date)).
- Re-upload CSV yang sama untuk tanggal yang sama akan mengganti nilai lama
  (bukan akumulasi), karena CSV adalah sumber kebenaran untuk periode itu.
"""

from collections import defaultdict
from datetime import date, datetime, timezone
from uuid import UUID

import sqlalchemy as sa
import sqlalchemy.dialects.postgresql as sa_pg
from fastapi import APIRouter, HTTPException, Query, UploadFile, status
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.csv_parser import (
    CsvParseError,
    MetaAdsRow,
    MetaBreakdownRow,
    ShopeeClickRow,
    ShopeeCommissionRow,
    is_live_platform,
    parse_meta_ads_csv,
    parse_meta_breakdown_csv,
    parse_shopee_click_csv,
    parse_shopee_commission_csv,
)
from app.core.deps import DB, CurrentUser
from app.models.account import MetaAccount, ShopeeAccount
from app.models.balance import AccountBalance
from app.models.campaign import Campaign, TagLink
from app.models.import_log import CsvImport
from app.models.metrics import ClickBySource, DailyMetric, MetaBreakdown, OrderSnapshot
from app.models.terra import TerraPlacement
from app.models.adu import AduPlacement
from app.schemas.upload import UploadResponse

router = APIRouter()


# ===========================================================================
# Endpoint publik
# ===========================================================================

@router.post("/meta-ads", response_model=UploadResponse, status_code=status.HTTP_201_CREATED)
async def upload_meta_ads(
    file: UploadFile,
    current_user: CurrentUser,
    db: DB,
    meta_account_id: UUID = Query(...),
):
    await _assert_meta_account_owned(meta_account_id, current_user.id, db)
    raw = await file.read()

    import_log = _start_import(current_user.id, "meta_ads", file.filename, db)

    try:
        rows = parse_meta_ads_csv(raw)
    except CsvParseError as exc:
        return await _fail_import(import_log, db, str(exc))

    # Catat spend lama untuk tanggal yang akan diupload (sebelum upsert)
    dates = {r.tanggal for r in rows}
    spend_lama = await _get_meta_spend_for_dates(meta_account_id, dates, db)

    ok, fail = 0, 0
    for row in rows:
        try:
            campaign = await _get_or_create_campaign(row, meta_account_id, db)
            await _upsert_meta_metric(campaign.id, row.tanggal, row.spend_idr, row.clicks_meta, db)
            ok += 1
        except Exception:
            fail += 1

    # Kurangi saldo: delta = spend baru (dari CSV) - spend lama (yang diganti)
    if ok > 0:
        spend_baru = sum(r.spend_idr for r in rows)
        await _adjust_balance_for_meta(meta_account_id, spend_baru - spend_lama, db)

    return await _finish_import(import_log, ok, fail, db)


@router.post("/shopee-commission", response_model=UploadResponse, status_code=status.HTTP_201_CREATED)
async def upload_shopee_commission(
    file: UploadFile,
    current_user: CurrentUser,
    db: DB,
    shopee_account_id: UUID = Query(...),
    tag_slot: int = Query(1, ge=1, le=5),
):
    await _assert_shopee_account_owned(shopee_account_id, current_user.id, db)
    raw = await file.read()

    import_log = _start_import(current_user.id, "shopee_commission", file.filename, db,
                               shopee_account_id=shopee_account_id)

    try:
        rows = parse_shopee_commission_csv(raw, tag_slot=tag_slot)
    except CsvParseError as exc:
        return await _fail_import(import_log, db, str(exc))

    ok, fail = 0, 0

    # Agregasi per (tag, tanggal) untuk DailyMetric
    grouped: dict[tuple[str, date], list[ShopeeCommissionRow]] = defaultdict(list)
    for row in rows:
        grouped[(row.tag, row.tanggal_order)].append(row)

    for (tag, tanggal), group in grouped.items():
        try:
            # SAVEPOINT per grup — kalau satu grup gagal (mis. constraint DB),
            # transaksi utama tidak ikut ke status aborted sehingga grup
            # berikutnya (dan flush snapshot di bawah) tetap bisa jalan.
            async with db.begin_nested():
                tag_link = await _get_or_create_tag_link(tag, shopee_account_id, db)
                selesai = sum(1 for r in group if r.status == "completed")
                # pending dan unpaid digabung ke tertunda (skema tidak pisahkan keduanya)
                tertunda = sum(1 for r in group if r.status in ("pending", "unpaid"))
                batal = sum(1 for r in group if r.status == "cancelled")
                commission = sum(r.commission_idr for r in group)
                commission_live = sum(r.commission_idr for r in group if is_live_platform(r.platform))
                sales = sum(r.sales_idr for r in group)
                await _upsert_commission_metric(
                    tag_link.id, tanggal, selesai, tertunda, batal, commission, commission_live, sales, db
                )
            ok += len(group)
        except Exception:
            fail += len(group)
            continue

    # Order snapshots — append-only per order_id
    for row in rows:
        try:
            async with db.begin_nested():
                tag_link = await _get_or_create_tag_link(row.tag, shopee_account_id, db)
                snapshot = OrderSnapshot(
                    shopee_account_id=shopee_account_id,
                    order_id=row.order_id,
                    tanggal_snapshot=row.tanggal_order,
                    status=row.status,
                    commission_from_idr=None,
                    commission_to_idr=row.commission_idr,
                    nama_produk=row.nama_produk,
                    nama_toko=row.nama_toko,
                    qty=row.qty,
                    sales_idr=row.sales_idr,
                )
                db.add(snapshot)
                await db.flush()
        except Exception:
            pass

    return await _finish_import(import_log, ok, fail, db)


@router.post("/meta-breakdown", response_model=UploadResponse, status_code=status.HTTP_201_CREATED)
async def upload_meta_breakdown(
    file: UploadFile,
    current_user: CurrentUser,
    db: DB,
    meta_account_id: UUID = Query(...),
):
    await _assert_meta_account_owned(meta_account_id, current_user.id, db)
    raw = await file.read()

    import_log = _start_import(current_user.id, "meta_breakdown", file.filename, db)

    try:
        rows = parse_meta_breakdown_csv(raw)
    except CsvParseError as exc:
        return await _fail_import(import_log, db, str(exc))

    ok, fail = 0, 0
    for row in rows:
        try:
            campaign = await _get_or_create_campaign(
                MetaAdsRow(
                    meta_campaign_id=row.meta_campaign_id,
                    nama_campaign=row.nama_campaign,
                    tanggal=row.tanggal,
                    spend_idr=row.spend_idr,
                    clicks_meta=row.clicks,
                ),
                meta_account_id,
                db,
            )
            stmt = sa_pg.insert(MetaBreakdown).values(
                campaign_id=campaign.id,
                tanggal=row.tanggal,
                tipe=row.tipe,
                nilai=row.nilai,
                spend_idr=row.spend_idr,
                impressions=row.impressions,
                clicks=row.clicks,
            )
            stmt = stmt.on_conflict_do_update(
                constraint="uq_meta_breakdown",
                set_={
                    "spend_idr": stmt.excluded.spend_idr,
                    "impressions": stmt.excluded.impressions,
                    "clicks": stmt.excluded.clicks,
                },
            )
            await db.execute(stmt)
            ok += 1
        except Exception:
            fail += 1

    return await _finish_import(import_log, ok, fail, db)


@router.post("/shopee-click", response_model=UploadResponse, status_code=status.HTTP_201_CREATED)
async def upload_shopee_click(
    file: UploadFile,
    current_user: CurrentUser,
    db: DB,
    shopee_account_id: UUID = Query(...),
    tag_slot: int = Query(1, ge=1, le=5),
):
    await _assert_shopee_account_owned(shopee_account_id, current_user.id, db)
    raw = await file.read()

    import_log = _start_import(current_user.id, "shopee_click", file.filename, db,
                               shopee_account_id=shopee_account_id)

    try:
        rows = parse_shopee_click_csv(raw)
    except CsvParseError as exc:
        return await _fail_import(import_log, db, str(exc))

    ok, fail = 0, 0

    # Agregasi per (tag, tanggal) untuk DailyMetric
    grouped: dict[tuple[str, date], int] = defaultdict(int)
    # Agregasi per (tag, tanggal, sumber) untuk click_by_source
    grouped_src: dict[tuple[str, date, str], int] = defaultdict(int)
    for row in rows:
        grouped[(row.tag, row.tanggal)] += row.clicks
        grouped_src[(row.tag, row.tanggal, row.sumber)] += row.clicks

    for (tag, tanggal), total_clicks in grouped.items():
        try:
            async with db.begin_nested():
                tag_link = await _get_or_create_tag_link(tag, shopee_account_id, db)
                await _upsert_click_metric(tag_link.id, tanggal, total_clicks, db)
            ok += 1
        except Exception:
            fail += 1

    for (tag, tanggal, sumber), clicks in grouped_src.items():
        try:
            async with db.begin_nested():
                tag_link = await _get_or_create_tag_link(tag, shopee_account_id, db)
                await _upsert_click_by_source(tag_link.id, tanggal, sumber, clicks, db)
        except Exception:
            pass

    return await _finish_import(import_log, ok, fail, db)


# ===========================================================================
# WD Payment — BillConversionReport Shopee
# ===========================================================================

@router.post("/wd-payment", response_model=UploadResponse, status_code=status.HTTP_201_CREATED)
async def upload_wd_payment(
    file: UploadFile,
    current_user: CurrentUser,
    db: DB,
    shopee_account_id: UUID = Query(...),
):
    """Upload BillConversionReport (.csv atau .xlsx) dari Shopee Affiliate."""
    from collections import defaultdict
    from datetime import timedelta, date as date_type
    from decimal import Decimal, InvalidOperation
    import io

    await _assert_shopee_account_owned(shopee_account_id, current_user.id, db)
    raw = await file.read()
    fname = (file.filename or "").lower()

    import_log = _start_import(current_user.id, "wd_payment", file.filename, db)

    # ── Parse file ────────────────────────────────────────────────────────────
    rows_raw: list[dict] = []
    try:
        if fname.endswith(".xlsx") or raw[:4] == b"PK\x03\x04":
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows_raw.append(dict(zip(headers, row)))
            wb.close()
        else:
            import csv
            text = raw.decode("utf-8-sig", errors="replace")
            # deteksi delimiter: jika baris pertama pakai ";" lebih banyak dari ","
            first_line = text.split("\n")[0]
            delim = ";" if first_line.count(";") > first_line.count(",") else ","
            reader = csv.DictReader(io.StringIO(text), delimiter=delim)
            for r in reader:
                rows_raw.append({k.strip(): v for k, v in r.items()})
    except Exception as exc:
        return await _fail_import(import_log, db, f"Gagal membaca file: {exc}")

    if not rows_raw:
        return await _fail_import(import_log, db, "File kosong atau tidak terbaca.")

    # ── Identifikasi kolom ────────────────────────────────────────────────────
    def _find_col(row: dict, *candidates):
        for c in candidates:
            if c in row:
                return c
        return None

    sample = rows_raw[0]
    col_status   = _find_col(sample,
        "Status Pesanan", "Status Pembelian", "Status Pemebelian", "Status")
    col_tgl      = _find_col(sample,
        "Waktu Terselesaikan", "Waktu Penyelesaian", "Tanggal Penyelesaian",
        "Waktu Pemesanan", "Tanggal Pesanan")
    col_komisi   = _find_col(sample,
        "Komisi Bersih Affiliate (Rp)", "Komisi Anda (Rp)",
        "Total Komisi per Pesanan(Rp)", "Komisi (Rp)")

    all_headers = list(sample.keys())
    if not all([col_status, col_tgl, col_komisi]):
        return await _fail_import(import_log, db,
            f"Kolom tidak ditemukan (status={col_status}, tgl={col_tgl}, komisi={col_komisi}). "
            f"Header file: {all_headers[:12]}")

    # ── Konversi tanggal (openpyxl bisa kembalikan datetime object) ───────────
    EXCEL_EPOCH = date_type(1899, 12, 30)

    def _parse_date(val) -> date_type | None:
        if val is None:
            return None
        # openpyxl mengembalikan datetime / date langsung
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date_type):
            return val
        if isinstance(val, (int, float)):
            try:
                return EXCEL_EPOCH + timedelta(days=float(val))
            except Exception:
                return None
        s = str(val).strip()
        # coba berbagai format string, termasuk yang ada komponen waktu
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S",
                    "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        try:
            return EXCEL_EPOCH + timedelta(days=float(s))
        except Exception:
            return None

    def _parse_idr(val) -> Decimal:
        """Parse angka IDR: handle titik ribuan ganda mis. '1.452.045' → 1452.045"""
        s = str(val or "0").strip()
        # ganti koma desimal ke titik
        s = s.replace(",", ".")
        parts = s.split(".")
        if len(parts) > 2:
            # titik pertama(s) = ribuan, titik terakhir = desimal
            s = "".join(parts[:-1]) + "." + parts[-1]
        return Decimal(s) if s else Decimal("0")

    # Status yang dianggap valid (selesai)
    STATUS_SELESAI = {"selesai", "completed", "complete", "sukses"}

    # ── Agregasi per tanggal ──────────────────────────────────────────────────
    agg: dict[date_type, dict] = defaultdict(lambda: {"komisi": Decimal("0"), "n": 0})
    status_sample: list[str] = []

    ok, fail = 0, 0
    for r in rows_raw:
        status_val = str(r.get(col_status) or "").strip()
        if len(status_sample) < 5 and status_val:
            status_sample.append(status_val)
        if status_val.lower() not in STATUS_SELESAI:
            continue
        tgl = _parse_date(r.get(col_tgl))
        if tgl is None:
            fail += 1
            continue
        try:
            k = _parse_idr(r.get(col_komisi))
        except (InvalidOperation, Exception):
            fail += 1
            continue
        agg[tgl]["komisi"] += k
        agg[tgl]["n"] += 1
        ok += 1

    if not agg:
        sample_info = f" Nilai status ditemukan: {list(set(status_sample))}" if status_sample else ""
        return await _fail_import(import_log, db,
            f"Tidak ada baris dengan status Selesai yang valid.{sample_info}")

    # ── Upsert ke wd_payments ─────────────────────────────────────────────────
    from app.models.wd_payment import WdPayment
    import uuid as uuid_mod

    for tgl, val in agg.items():
        stmt = sa_pg.insert(WdPayment).values(
            id=uuid_mod.uuid4(),
            shopee_account_id=shopee_account_id,
            user_id=current_user.id,
            tanggal=tgl,
            total_komisi=val["komisi"],
            jumlah_orders=val["n"],
        ).on_conflict_do_update(
            constraint="uq_wd_payments_account_tanggal",
            set_={"total_komisi": val["komisi"], "jumlah_orders": val["n"],
                  "updated_at": sa.text("now()")},
        )
        await db.execute(stmt)

    await db.commit()
    return await _finish_import(import_log, ok, fail, db)


@router.get("/wd-payment")
async def get_wd_payments(
    current_user: CurrentUser,
    db: DB,
    tanggal_dari: date = Query(...),
    tanggal_sampai: date = Query(...),
    shopee_account_id: UUID | None = Query(None),
):
    """Ambil data WD payment per tanggal untuk Komisi Bersih."""
    from app.models.wd_payment import WdPayment

    q = (
        sa.select(WdPayment.tanggal, sa.func.sum(WdPayment.total_komisi).label("total_komisi"),
                  sa.func.sum(WdPayment.jumlah_orders).label("jumlah_orders"))
        .join(ShopeeAccount, WdPayment.shopee_account_id == ShopeeAccount.id)
        .where(
            ShopeeAccount.user_id == current_user.id,
            WdPayment.tanggal.between(tanggal_dari, tanggal_sampai),
        )
        .group_by(WdPayment.tanggal)
        .order_by(WdPayment.tanggal)
    )
    if shopee_account_id:
        q = q.where(WdPayment.shopee_account_id == shopee_account_id)

    rows = (await db.execute(q)).all()
    return [{"tanggal": str(r.tanggal), "total_komisi": float(r.total_komisi),
             "jumlah_orders": r.jumlah_orders} for r in rows]


# ===========================================================================
# Terra Ads — placement spend dengan konversi USD → IDR
# ===========================================================================

_TERRA_KURS = 19_000  # 1 USD = Rp 19.000


@router.post("/terra", response_model=UploadResponse, status_code=status.HTTP_201_CREATED)
async def upload_terra(
    file: UploadFile,
    current_user: CurrentUser,
    db: DB,
    tanggal: date = Query(..., description="Tanggal data dalam format YYYY-MM-DD"),
):
    """Upload Terra Ads placement CSV. Kolom Spent (USD) dikonversi ke budget_rupiah × 19.000."""
    import csv as csv_mod
    import io
    from decimal import Decimal, InvalidOperation

    raw = await file.read()
    import_log = _start_import(current_user.id, "terra", file.filename, db)

    try:
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv_mod.DictReader(io.StringIO(text))
        if "Spent" not in (reader.fieldnames or []):
            return await _fail_import(import_log, db,
                "Kolom 'Spent' tidak ditemukan. Pastikan file adalah Terra Ads placement CSV.")
    except Exception as exc:
        return await _fail_import(import_log, db, f"Gagal membaca file: {exc}")

    ok, fail = 0, 0
    for row in reader:
        placement_id = (row.get("Placement") or "").strip()
        if not placement_id:
            fail += 1
            continue
        spent_raw = (row.get("Spent") or "").strip().lstrip("$").replace(",", "")
        if not spent_raw or spent_raw == "-":
            fail += 1
            continue
        try:
            spent_usd = Decimal(spent_raw)
        except InvalidOperation:
            fail += 1
            continue

        budget_rupiah = round(spent_usd * _TERRA_KURS)
        state = (row.get("State") or "").strip()
        try:
            impressions = int((row.get("Impressions") or "0").replace(",", ""))
        except (ValueError, AttributeError):
            impressions = 0
        try:
            clicks = int((row.get("Clicks") or "0").replace(",", ""))
        except (ValueError, AttributeError):
            clicks = 0

        try:
            stmt = sa_pg.insert(TerraPlacement).values(
                user_id=current_user.id,
                tanggal=tanggal,
                placement_id=placement_id,
                state=state,
                impressions=impressions,
                clicks=clicks,
                spent_usd=spent_usd,
                budget_rupiah=budget_rupiah,
            ).on_conflict_do_update(
                constraint="uq_terra_placement",
                set_={
                    "state": state,
                    "impressions": impressions,
                    "clicks": clicks,
                    "spent_usd": spent_usd,
                    "budget_rupiah": budget_rupiah,
                    "uploaded_at": sa.text("now()"),
                },
            )
            await db.execute(stmt)
            ok += 1
        except Exception:
            fail += 1

    return await _finish_import(import_log, ok, fail, db)


# ===========================================================================
# Adu Ads — zone placement spend dengan konversi USD → IDR
# ===========================================================================

_ADU_KURS = 19_000  # 1 USD = Rp 19.000


@router.post("/adu", response_model=UploadResponse, status_code=status.HTTP_201_CREATED)
async def upload_adu(
    file: UploadFile,
    current_user: CurrentUser,
    db: DB,
    tanggal: date = Query(..., description="Tanggal data dalam format YYYY-MM-DD"),
):
    """Upload Adu Ads zone CSV. Kolom Cost (USD) dikonversi ke budget_rupiah × 19.000."""
    import csv as csv_mod
    import io
    from decimal import Decimal, InvalidOperation

    raw = await file.read()
    import_log = _start_import(current_user.id, "adu", file.filename, db)

    try:
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv_mod.DictReader(io.StringIO(text))
        # normalisasi nama kolom (strip whitespace dan kutip)
        fieldnames = [f.strip().strip('"') for f in (reader.fieldnames or [])]
        if "Cost" not in fieldnames and "cost" not in [f.lower() for f in fieldnames]:
            return await _fail_import(import_log, db,
                "Kolom 'Cost' tidak ditemukan. Pastikan file adalah Adu Ads zone CSV.")
    except Exception as exc:
        return await _fail_import(import_log, db, f"Gagal membaca file: {exc}")

    def _col(row: dict, name: str) -> str:
        """Ambil nilai kolom case-insensitive, strip whitespace dan kutip."""
        for k, v in row.items():
            if k.strip().strip('"').lower() == name.lower():
                return str(v or "").strip().strip('"').strip()
        return ""

    def _int(s: str) -> int:
        try:
            return int(s.replace(",", "").split(".")[0])
        except (ValueError, AttributeError):
            return 0

    ok, fail = 0, 0
    for row in reader:
        zone_id = _col(row, "Zone ID")
        if not zone_id:
            fail += 1
            continue
        cost_raw = _col(row, "Cost")
        if not cost_raw or cost_raw == "0":
            # baris cost 0 tetap disimpan agar data lengkap
            cost_raw = "0"
        try:
            cost_usd = Decimal(cost_raw)
        except InvalidOperation:
            fail += 1
            continue

        budget_rupiah = round(cost_usd * _ADU_KURS)

        try:
            stmt = sa_pg.insert(AduPlacement).values(
                user_id=current_user.id,
                tanggal=tanggal,
                zone_id=zone_id,
                impressions=_int(_col(row, "Impressions")),
                clicks=_int(_col(row, "Clicks")),
                conversions=_int(_col(row, "Conversions")),
                cost_usd=cost_usd,
                budget_rupiah=budget_rupiah,
            ).on_conflict_do_update(
                constraint="uq_adu_placement",
                set_={
                    "impressions": sa.text("excluded.impressions"),
                    "clicks": sa.text("excluded.clicks"),
                    "conversions": sa.text("excluded.conversions"),
                    "cost_usd": sa.text("excluded.cost_usd"),
                    "budget_rupiah": sa.text("excluded.budget_rupiah"),
                    "uploaded_at": sa.text("now()"),
                },
            )
            await db.execute(stmt)
            ok += 1
        except Exception:
            fail += 1

    return await _finish_import(import_log, ok, fail, db)


# ===========================================================================
# Upsert helpers — masing-masing hanya update kolom miliknya
# ===========================================================================

async def _upsert_meta_metric(
    campaign_id: UUID, tanggal: date, spend_idr, clicks_meta: int, db: AsyncSession
) -> None:
    stmt = pg_insert(DailyMetric).values(
        campaign_id=campaign_id,
        tanggal=tanggal,
        spend_idr=spend_idr,
        clicks_meta=clicks_meta,
    )
    stmt = stmt.on_conflict_do_update(
        index_elements=["campaign_id", "tanggal"],
        index_where=sa.text("campaign_id IS NOT NULL"),
        set_={"spend_idr": stmt.excluded.spend_idr, "clicks_meta": stmt.excluded.clicks_meta},
    )
    await db.execute(stmt)


async def _upsert_commission_metric(
    tag_link_id: UUID, tanggal: date,
    selesai: int, tertunda: int, batal: int, commission_idr, commission_live_idr, sales_idr,
    db: AsyncSession,
) -> None:
    stmt = pg_insert(DailyMetric).values(
        tag_link_id=tag_link_id,
        tanggal=tanggal,
        orders_selesai=selesai,
        orders_tertunda=tertunda,
        orders_batal=batal,
        commission_idr=commission_idr,
        commission_live_idr=commission_live_idr,
        sales_idr=sales_idr,
    )
    stmt = stmt.on_conflict_do_update(
        index_elements=["tag_link_id", "tanggal"],
        index_where=sa.text("tag_link_id IS NOT NULL"),
        set_={
            "orders_selesai": stmt.excluded.orders_selesai,
            "orders_tertunda": stmt.excluded.orders_tertunda,
            "orders_batal": stmt.excluded.orders_batal,
            "commission_idr": stmt.excluded.commission_idr,
            "commission_live_idr": stmt.excluded.commission_live_idr,
            "sales_idr": stmt.excluded.sales_idr,
        },
    )
    await db.execute(stmt)


async def _upsert_click_metric(
    tag_link_id: UUID, tanggal: date, clicks_shopee: int, db: AsyncSession
) -> None:
    stmt = pg_insert(DailyMetric).values(
        tag_link_id=tag_link_id,
        tanggal=tanggal,
        clicks_shopee=clicks_shopee,
    )
    stmt = stmt.on_conflict_do_update(
        index_elements=["tag_link_id", "tanggal"],
        index_where=sa.text("tag_link_id IS NOT NULL"),
        set_={"clicks_shopee": stmt.excluded.clicks_shopee},
    )
    await db.execute(stmt)


async def _upsert_click_by_source(
    tag_link_id: UUID, tanggal: date, sumber: str, clicks: int, db: AsyncSession
) -> None:
    stmt = pg_insert(ClickBySource).values(
        tag_link_id=tag_link_id,
        tanggal=tanggal,
        sumber=sumber,
        clicks=clicks,
    )
    stmt = stmt.on_conflict_do_update(
        constraint="uq_click_by_source",
        set_={"clicks": stmt.excluded.clicks},
    )
    await db.execute(stmt)


# ===========================================================================
# Find-or-create: Campaign & TagLink
# ===========================================================================

async def _get_or_create_campaign(row: MetaAdsRow, meta_account_id: UUID, db: AsyncSession) -> Campaign:
    result = await db.execute(
        sa.select(Campaign).where(
            Campaign.meta_account_id == meta_account_id,
            Campaign.meta_campaign_id == row.meta_campaign_id,
        )
    )
    campaign = result.scalar_one_or_none()
    if campaign:
        # Update nama kalau berubah di Meta
        if campaign.nama_campaign != row.nama_campaign:
            campaign.nama_campaign = row.nama_campaign
        return campaign

    campaign = Campaign(
        meta_account_id=meta_account_id,
        meta_campaign_id=row.meta_campaign_id,
        nama_campaign=row.nama_campaign,
    )
    db.add(campaign)
    await db.flush()  # dapat ID sebelum dipakai di upsert
    return campaign


async def _get_or_create_tag_link(tag: str, shopee_account_id: UUID, db: AsyncSession) -> TagLink:
    result = await db.execute(
        sa.select(TagLink).where(
            TagLink.shopee_account_id == shopee_account_id,
            TagLink.tag == tag,
        )
    )
    tag_link = result.scalar_one_or_none()
    if tag_link:
        return tag_link

    tag_link = TagLink(shopee_account_id=shopee_account_id, tag=tag)
    db.add(tag_link)
    await db.flush()
    return tag_link


# ===========================================================================
# Import log helpers
# ===========================================================================

def _start_import(
    user_id, tipe: str, nama_file: str, db: AsyncSession,
    shopee_account_id: UUID | None = None,
) -> CsvImport:
    log = CsvImport(
        user_id=user_id,
        shopee_account_id=shopee_account_id,
        tipe=tipe,
        nama_file=nama_file or "unknown.csv",
        file_ref=nama_file or "unknown.csv",   # MVP: simpan nama file saja, belum ada object storage
        status="diproses",
    )
    db.add(log)
    return log


async def _finish_import(log: CsvImport, ok: int, fail: int, db: AsyncSession) -> CsvImport:
    log.baris_diproses = ok
    log.baris_gagal = fail
    log.status = "selesai"
    log.processed_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(log)
    return log


async def _fail_import(log: CsvImport, db: AsyncSession, detail: str) -> None:
    log.status = "gagal"
    log.catatan = detail
    log.processed_at = datetime.now(timezone.utc)
    await db.commit()
    raise HTTPException(status_code=422, detail=detail)


# ===========================================================================
# Guard: pastikan akun milik user yang request
# ===========================================================================

async def _assert_meta_account_owned(account_id: UUID, user_id, db: AsyncSession) -> None:
    result = await db.execute(
        sa.select(MetaAccount).where(MetaAccount.id == account_id, MetaAccount.user_id == user_id)
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Akun Meta tidak ditemukan.")


async def _assert_shopee_account_owned(account_id: UUID, user_id, db: AsyncSession) -> None:
    result = await db.execute(
        sa.select(ShopeeAccount).where(ShopeeAccount.id == account_id, ShopeeAccount.user_id == user_id)
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Akun Shopee tidak ditemukan.")


# ===========================================================================
# Saldo helpers
# ===========================================================================

async def _get_meta_spend_for_dates(
    meta_account_id: UUID, dates: set, db: AsyncSession
) -> "Decimal":
    from decimal import Decimal
    if not dates:
        return Decimal("0")
    result = await db.execute(
        sa.select(sa.func.coalesce(sa.func.sum(DailyMetric.spend_idr), 0))
        .join(Campaign, DailyMetric.campaign_id == Campaign.id)
        .where(
            Campaign.meta_account_id == meta_account_id,
            DailyMetric.campaign_id.isnot(None),
            DailyMetric.tanggal.in_(dates),
        )
    )
    return result.scalar() or Decimal("0")


async def _adjust_balance_for_meta(
    meta_account_id: UUID, delta: "Decimal", db: AsyncSession
) -> None:
    """Kurangi sisa_saldo sebesar delta (spend baru - spend lama). Tidak pernah di bawah 0."""
    from decimal import Decimal
    if not delta or delta <= 0:
        return
    bal = (await db.execute(
        sa.select(AccountBalance).where(AccountBalance.meta_account_id == meta_account_id)
    )).scalar_one_or_none()
    if bal:
        bal.sisa_saldo = max(Decimal("0"), bal.sisa_saldo - delta)
